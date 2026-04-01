import io
import json

from flask import Blueprint, render_template, request, redirect, url_for, jsonify
from openpyxl import load_workbook
from sqlalchemy import or_, func
from urllib.parse import urlencode

from models import (
    db,
    DonVi,
    DanhMuc,
    DanhMucField,
    DanhMucRecord,
    DanhMucRecordValue,
    DanhMucDataset
)

from services.bhyt_service import (
    check_bhyt_root_status,
    refresh_bhyt_captcha,
    login_bhyt,
    post_sync_export,
    get_sync_stream,
    clear_dataset_records,
    validate_sync_config,
)

category_bp = Blueprint("category_bp", __name__, url_prefix="/categories")


@category_bp.route("/", methods=["GET"])
def list_categories():
    keyword = (request.args.get("keyword") or "").strip()
    scope = (request.args.get("scope") or "").strip().upper()

    query = DanhMuc.query

    if keyword:
        query = query.filter(
            or_(
                DanhMuc.ten_danh_muc.ilike(f"%{keyword}%")
            )
        )

    if scope in ["COMMON", "UNIT"]:
        query = query.filter(DanhMuc.scope == scope)

    items = query.order_by(DanhMuc.id.asc()).all()

    common_items = [x for x in items if (x.scope or "").upper() == "COMMON"]
    unit_items = [x for x in items if (x.scope or "").upper() == "UNIT"]

    return render_template(
        "categories.html",
        items=items,
        common_items=common_items,
        unit_items=unit_items,
        keyword=keyword,
        scope=scope
    )


@category_bp.route("/create", methods=["GET", "POST"])
def create_category():
    error = None

    if request.method == "POST":
        try:
            ten_danh_muc = (request.form.get("ten_danh_muc") or "").strip()
            scope = (request.form.get("scope") or "COMMON").strip().upper()
            sync_don_vi_id = to_int_or_none(request.form.get("sync_don_vi_id"))
            api_danh_muc_url = (request.form.get("api_danh_muc_url") or "").strip()
            api_tim_kiem_url = (request.form.get("api_tim_kiem_url") or "").strip()
            api_tim_kiem_body = (request.form.get("api_tim_kiem_body") or "").strip()
            api_tong_hop_url = (request.form.get("api_tong_hop_url") or "").strip()
            api_xuat_file_url = (request.form.get("api_xuat_file_url") or "").strip()

            if not ten_danh_muc:
                raise ValueError("Vui lòng nhập tên danh mục.")

            if scope not in ["COMMON", "UNIT"]:
                raise ValueError("Phạm vi danh mục không hợp lệ.")

            item = DanhMuc(
                ten_danh_muc=ten_danh_muc,
                scope=scope,
                sync_don_vi_id=sync_don_vi_id if scope == "COMMON" else None,
                api_danh_muc_url=api_danh_muc_url or None,
                api_tim_kiem_url=api_tim_kiem_url or None,
                api_tim_kiem_body=api_tim_kiem_body or None,
                api_tong_hop_url=api_tong_hop_url or None,
                api_xuat_file_url=api_xuat_file_url or None
            )
            db.session.add(item)
            db.session.flush()

            if scope == "COMMON":
                db.session.add(DanhMucDataset(
                    danh_muc_id=item.id,
                    don_vi_id=None,
                    ten_bo_du_lieu="Bộ dữ liệu chung"
                ))

            db.session.commit()
            return redirect(url_for("category_bp.list_categories"))

        except Exception as e:
            db.session.rollback()
            error = str(e)

    return render_template(
        "category_form.html",
        item=None,
        units=DonVi.query.order_by(DonVi.id.asc()).all(),
        error=error
    )


@category_bp.route("/<int:category_id>/edit", methods=["GET", "POST"])
def edit_category(category_id):
    item = DanhMuc.query.get_or_404(category_id)
    error = None

    if request.method == "POST":
        try:
            ten_danh_muc = (request.form.get("ten_danh_muc") or "").strip()
            scope = (request.form.get("scope") or "COMMON").strip().upper()
            sync_don_vi_id = to_int_or_none(request.form.get("sync_don_vi_id"))
            api_danh_muc_url = (request.form.get("api_danh_muc_url") or "").strip()
            api_tim_kiem_url = (request.form.get("api_tim_kiem_url") or "").strip()
            api_tim_kiem_body = (request.form.get("api_tim_kiem_body") or "").strip()
            api_tong_hop_url = (request.form.get("api_tong_hop_url") or "").strip()
            api_xuat_file_url = (request.form.get("api_xuat_file_url") or "").strip()

            if not ten_danh_muc:
                raise ValueError("Vui lòng nhập tên danh mục.")

            if scope not in ["COMMON", "UNIT"]:
                raise ValueError("Phạm vi danh mục không hợp lệ.")

            old_scope = (item.scope or "COMMON").upper()

            if old_scope == "COMMON" and scope == "UNIT":
                common_dataset = (
                    DanhMucDataset.query
                    .filter(DanhMucDataset.danh_muc_id == item.id)
                    .filter(DanhMucDataset.don_vi_id.is_(None))
                    .first()
                )
                if common_dataset:
                    has_records = DanhMucRecord.query.filter_by(dataset_id=common_dataset.id).first()
                    if has_records:
                        raise ValueError(
                            "Danh mục đang có dữ liệu chung. Vui lòng xóa/chuyển dữ liệu trước khi đổi sang danh mục riêng."
                        )

            item.ten_danh_muc = ten_danh_muc
            item.scope = scope
            item.sync_don_vi_id = sync_don_vi_id if scope == "COMMON" else None
            item.api_danh_muc_url = api_danh_muc_url or None
            item.api_tim_kiem_url = api_tim_kiem_url or None
            item.api_tim_kiem_body = api_tim_kiem_body or None
            item.api_tong_hop_url = api_tong_hop_url or None
            item.api_xuat_file_url = api_xuat_file_url or None

            if scope == "COMMON":
                get_or_create_common_dataset(item)

            db.session.commit()
            return redirect(url_for("category_bp.list_categories"))

        except Exception as e:
            db.session.rollback()
            error = str(e)

    return render_template(
        "category_form.html",
        item=item,
        units=DonVi.query.order_by(DonVi.id.asc()).all(),
        error=error
    )


@category_bp.route("/<int:category_id>/delete", methods=["POST"])
def delete_category(category_id):
    item = DanhMuc.query.get_or_404(category_id)
    db.session.delete(item)
    db.session.commit()
    return redirect(url_for("category_bp.list_categories"))


@category_bp.route("/api/fields", methods=["GET"])
def get_category_fields_api():
    category_id = request.args.get("category_id", type=int)

    if not category_id:
        return jsonify([])

    fields = (
        DanhMucField.query
        .filter_by(danh_muc_id=category_id)
        .order_by(DanhMucField.id.asc())
        .all()
    )

    result = []
    for f in fields:
        result.append({
            "id": f.id,
            "ma_truong": f.ma_truong
        })

    return jsonify(result)


@category_bp.route("/<int:category_id>/fields", methods=["GET", "POST"])
def manage_category_fields(category_id):
    item = DanhMuc.query.get_or_404(category_id)
    error = None

    if request.method == "POST":
        action = (request.form.get("action") or "").strip()

        try:
            if action == "create":
                ma_truong = (request.form.get("ma_truong") or "").strip()

                if not ma_truong:
                    raise ValueError("Vui lòng nhập mã trường.")

                existed = DanhMucField.query.filter_by(
                    danh_muc_id=item.id,
                    ma_truong=ma_truong
                ).first()
                if existed:
                    raise ValueError("Mã trường đã tồn tại trong danh mục này.")

                db.session.add(DanhMucField(
                    danh_muc_id=item.id,
                    ma_truong=ma_truong
                ))
                db.session.commit()

                return redirect(url_for("category_bp.manage_category_fields", category_id=item.id))

            if action == "delete":
                field_id = to_int_or_none(request.form.get("field_id"))
                field = DanhMucField.query.filter_by(id=field_id, danh_muc_id=item.id).first()
                if not field:
                    raise ValueError("Không tìm thấy field.")

                DanhMucRecordValue.query.filter_by(field_id=field.id).delete()
                db.session.delete(field)
                db.session.commit()

                return redirect(url_for("category_bp.manage_category_fields", category_id=item.id))

            raise ValueError("Thao tác không hợp lệ.")

        except Exception as e:
            db.session.rollback()
            error = str(e)

    fields = (
        DanhMucField.query
        .filter_by(danh_muc_id=item.id)
        .order_by(DanhMucField.id.asc())
        .all()
    )

    return render_template(
        "category_fields.html",
        item=item,
        fields=fields,
        error=error
    )


@category_bp.route("/<int:category_id>/datasets", methods=["GET", "POST"])
def manage_category_datasets(category_id):
    item = DanhMuc.query.get_or_404(category_id)
    error = None

    if (item.scope or "COMMON").upper() != "UNIT":
        return redirect(url_for("category_bp.manage_category_records", category_id=item.id))

    units = DonVi.query.order_by(DonVi.id.asc()).all()

    if request.method == "POST":
        action = (request.form.get("action") or "").strip()

        try:
            if action != "create":
                raise ValueError("Thao tác không hợp lệ.")

            don_vi_id = to_int_or_none(request.form.get("don_vi_id"))
            ten_bo_du_lieu = (request.form.get("ten_bo_du_lieu") or "").strip()

            if not don_vi_id:
                raise ValueError("Vui lòng chọn đơn vị.")

            unit = DonVi.query.get(don_vi_id)
            if not unit:
                raise ValueError("Đơn vị không tồn tại.")

            existed = (
                DanhMucDataset.query
                .filter_by(danh_muc_id=item.id, don_vi_id=don_vi_id)
                .first()
            )
            if existed:
                raise ValueError("Đơn vị này đã có dữ liệu.")

            if not ten_bo_du_lieu:
                ten_bo_du_lieu = f"Bộ dữ liệu - {unit.ten_don_vi}"

            db.session.add(DanhMucDataset(
                danh_muc_id=item.id,
                don_vi_id=don_vi_id,
                ten_bo_du_lieu=ten_bo_du_lieu
            ))
            db.session.commit()

            return redirect(url_for("category_bp.manage_category_datasets", category_id=item.id))

        except Exception as e:
            db.session.rollback()
            error = str(e)

    datasets = (
        DanhMucDataset.query
        .filter_by(danh_muc_id=item.id)
        .order_by(DanhMucDataset.id.asc())
        .all()
    )

    return render_template(
        "category_datasets.html",
        item=item,
        datasets=datasets,
        units=units,
        error=error
    )


@category_bp.route("/<int:category_id>/datasets/<int:dataset_id>/delete", methods=["POST"])
def delete_category_dataset(category_id, dataset_id):
    item = DanhMuc.query.get_or_404(category_id)

    dataset = (
        DanhMucDataset.query
        .filter_by(id=dataset_id, danh_muc_id=item.id)
        .first_or_404()
    )

    db.session.delete(dataset)
    db.session.commit()

    return redirect(url_for("category_bp.manage_category_datasets", category_id=item.id))


@category_bp.route("/<int:category_id>/records", methods=["GET", "POST"])
def manage_category_records(category_id):
    item = DanhMuc.query.get_or_404(category_id)

    if (item.scope or "COMMON").upper() == "UNIT":
        return redirect(url_for("category_bp.manage_category_datasets", category_id=item.id))

    dataset = get_or_create_common_dataset(item)
    db.session.commit()

    return handle_dataset_records(item=item, dataset=dataset)


@category_bp.route("/<int:category_id>/datasets/<int:dataset_id>/records", methods=["GET", "POST"])
def manage_category_dataset_records(category_id, dataset_id):
    item = DanhMuc.query.get_or_404(category_id)

    dataset = (
        DanhMucDataset.query
        .filter_by(id=dataset_id, danh_muc_id=item.id)
        .first_or_404()
    )

    return handle_dataset_records(item=item, dataset=dataset)


@category_bp.route("/<int:category_id>/datasets/<int:dataset_id>/records/clear", methods=["POST"])
def clear_category_dataset_records(category_id, dataset_id):
    item = DanhMuc.query.get_or_404(category_id)

    dataset = (
        DanhMucDataset.query
        .filter_by(id=dataset_id, danh_muc_id=item.id)
        .first_or_404()
    )

    try:
        clear_dataset_records(dataset)
        db.session.commit()

        return redirect(
            url_for(
                "category_bp.manage_category_dataset_records",
                category_id=item.id,
                dataset_id=dataset.id
            )
        )
    except Exception as e:
        db.session.rollback()
        return redirect(
            url_for(
                "category_bp.manage_category_dataset_records",
                category_id=item.id,
                dataset_id=dataset.id,
                error=str(e)
            )
        )


@category_bp.route("/<int:category_id>/records/clear", methods=["POST"])
def clear_category_records(category_id):
    item = DanhMuc.query.get_or_404(category_id)

    if (item.scope or "COMMON").upper() == "UNIT":
        return redirect(url_for("category_bp.manage_category_datasets", category_id=item.id))

    dataset = get_or_create_common_dataset(item)

    try:
        clear_dataset_records(dataset)
        db.session.commit()

        return redirect(
            url_for(
                "category_bp.manage_category_records",
                category_id=item.id
            )
        )
    except Exception as e:
        db.session.rollback()
        return redirect(
            url_for(
                "category_bp.manage_category_records",
                category_id=item.id,
                error=str(e)
            )
        )


@category_bp.route("/<int:category_id>/sync", methods=["POST"])
def sync_category_common(category_id):
    item = DanhMuc.query.get_or_404(category_id)

    if (item.scope or "COMMON").upper() != "COMMON":
        return jsonify({"error": "Route này chỉ dùng cho danh mục chung."}), 400

    dataset = get_or_create_common_dataset(item)
    return run_sync_for_category(item=item, dataset=dataset, unit=resolve_sync_unit(item=item, dataset=dataset))


@category_bp.route("/<int:category_id>/sync-login", methods=["POST"])
def sync_category_common_login(category_id):
    item = DanhMuc.query.get_or_404(category_id)

    if (item.scope or "COMMON").upper() != "COMMON":
        return jsonify({"error": "Route này chỉ dùng cho danh mục chung."}), 400

    return run_sync_login_for_unit(resolve_sync_unit(item=item, dataset=None))


@category_bp.route("/<int:category_id>/datasets/<int:dataset_id>/sync", methods=["POST"])
def sync_category_dataset(category_id, dataset_id):
    item = DanhMuc.query.get_or_404(category_id)

    dataset = (
        DanhMucDataset.query
        .filter_by(id=dataset_id, danh_muc_id=item.id)
        .first_or_404()
    )

    return run_sync_for_category(item=item, dataset=dataset, unit=resolve_sync_unit(item=item, dataset=dataset))


@category_bp.route("/<int:category_id>/datasets/<int:dataset_id>/sync-login", methods=["POST"])
def sync_category_dataset_login(category_id, dataset_id):
    item = DanhMuc.query.get_or_404(category_id)

    dataset = (
        DanhMucDataset.query
        .filter_by(id=dataset_id, danh_muc_id=item.id)
        .first_or_404()
    )

    return run_sync_login_for_unit(resolve_sync_unit(item=item, dataset=dataset))


def run_sync_for_category(item, dataset, unit):
    try:
        validate_sync_config(item)

        if not unit:
            raise ValueError("Chưa xác định được đơn vị để đồng bộ.")

        root_resp = check_bhyt_root_status(unit)

        if root_resp.status_code == 200:
            captcha_html = refresh_bhyt_captcha(unit)
            return jsonify({
                "need_login": True,
                "captcha_html": captcha_html
            })

        if root_resp.status_code != 302:
            raise ValueError(f"Không kiểm tra được trạng thái đăng nhập BHYT. HTTP {root_resp.status_code}")

        post_result = post_sync_export(unit, item)
        post_data = post_result["data"]

        if str(post_data.get("result")).upper() != "OK":
            raise ValueError(f"API tổng hợp Excel phản hồi không hợp lệ: {post_data}")

        file_bytes = get_sync_stream(unit, item.api_xuat_file_url)

        clear_dataset_records(dataset)
        db.session.flush()

        DanhMucRecordValue.query.filter(
            DanhMucRecordValue.field_id.in_(
                db.session.query(DanhMucField.id).filter(DanhMucField.danh_muc_id == item.id)
            )
        ).delete(synchronize_session=False)
        DanhMucField.query.filter_by(danh_muc_id=item.id).delete(synchronize_session=False)
        db.session.flush()

        import_excel_rows(dataset, [], file_bytes, auto_create_fields=True)
        db.session.commit()

        redirect_url = build_dataset_records_url(item=item, dataset=dataset, page=1, page_size=50)

        return jsonify({
            "success": True,
            "message": f"Đồng bộ thành công. Số bản ghi export: {post_data.get('countList', 'không rõ')}",
            "redirect_url": redirect_url
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 400


def run_sync_login_for_unit(unit):
    if not unit:
        return jsonify({"error": "Chưa xác định được đơn vị để đăng nhập đồng bộ."}), 400

    try:
        payload = request.get_json(silent=True) or {}
        captcha = (payload.get("captcha") or "").strip()

        if not captcha:
            raise ValueError("Vui lòng nhập captcha.")

        login_result = login_bhyt(unit, captcha)
        data = login_result["data"]

        if data.get("data") != 1:
            if data.get("data") == -1:
                raise ValueError("Sai tài khoản hoặc mật khẩu cổng BHYT.")
            if data.get("data") == -2:
                raise ValueError("Sai captcha.")
            raise ValueError(f"Đăng nhập BHYT thất bại: {data}")

        return jsonify({
            "success": True,
            "message": "Đăng nhập BHYT thành công."
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 400


def resolve_sync_unit(item, dataset=None):
    scope = (item.scope or "COMMON").upper()

    if scope == "UNIT":
        if dataset and dataset.don_vi:
            return dataset.don_vi
        raise ValueError("Bộ dữ liệu chưa gắn đơn vị.")

    if item.sync_don_vi:
        return item.sync_don_vi

    if item.sync_don_vi_id:
        return DonVi.query.get(item.sync_don_vi_id)

    raise ValueError("Danh mục chung chưa chọn đơn vị gốc để đồng bộ.")


def handle_dataset_records(item, dataset):
    fields = (
        DanhMucField.query
        .filter_by(danh_muc_id=item.id)
        .order_by(DanhMucField.id.asc())
        .all()
    )
    error = request.args.get("error") or None
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"

    page_size = normalize_page_size(request.args.get("page_size", type=int, default=50))
    page = request.args.get("page", type=int, default=1) or 1
    filter_conditions = parse_filter_conditions(request.args)

    if request.method == "POST":
        action = (request.form.get("action") or "").strip()

        try:
            if not fields and action in ["create_manual", "import_json", "import_excel", "update_record"]:
                raise ValueError("Danh mục chưa có field, vui lòng tạo field trước.")

            if action == "create_manual":
                record = DanhMucRecord(dataset_id=dataset.id)
                db.session.add(record)
                db.session.flush()

                cell_values = []
                for field in fields:
                    val = (request.form.get(f"field_{field.id}") or "").strip()
                    norm_val = val or None
                    db.session.add(DanhMucRecordValue(
                        record_id=record.id,
                        field_id=field.id,
                        value=norm_val
                    ))
                    cell_values.append(val)

                db.session.commit()

                redirect_url = build_dataset_records_url(item=item, dataset=dataset, page=1, page_size=page_size)
                if is_ajax:
                    return jsonify({
                        "success": True,
                        "record_id": record.id,
                        "cells": cell_values,
                        "redirect_url": redirect_url
                    })

                return redirect(redirect_url)

            if action == "update_record":
                record_id = to_int_or_none(request.form.get("record_id"))
                record = DanhMucRecord.query.filter_by(id=record_id, dataset_id=dataset.id).first()
                if not record:
                    raise ValueError("Không tìm thấy dòng dữ liệu cần sửa.")

                existing_values = {v.field_id: v for v in record.values}
                cell_values = []

                for field in fields:
                    val = (request.form.get(f"field_{field.id}") or "").strip()
                    norm_val = val or None
                    value_obj = existing_values.get(field.id)

                    if value_obj:
                        value_obj.value = norm_val
                    else:
                        db.session.add(DanhMucRecordValue(
                            record_id=record.id,
                            field_id=field.id,
                            value=norm_val
                        ))

                    cell_values.append(val)

                db.session.commit()

                if is_ajax:
                    return jsonify({
                        "success": True,
                        "record_id": record.id,
                        "cells": cell_values
                    })

                return redirect(build_dataset_records_url(item=item, dataset=dataset, page=page, page_size=page_size, filter_conditions=filter_conditions))

            if action == "delete_record":
                record_id = to_int_or_none(request.form.get("record_id"))
                record = DanhMucRecord.query.filter_by(id=record_id, dataset_id=dataset.id).first()
                if not record:
                    raise ValueError("Không tìm thấy dòng dữ liệu.")

                DanhMucRecordValue.query.filter_by(record_id=record.id).delete()
                db.session.delete(record)
                db.session.commit()

                total_after_delete = build_records_query(dataset.id, filter_conditions).count()
                total_pages_after_delete = max(1, (total_after_delete + page_size - 1) // page_size)
                new_page = min(page, total_pages_after_delete)
                redirect_url = build_dataset_records_url(item=item, dataset=dataset, page=new_page, page_size=page_size, filter_conditions=filter_conditions)

                if is_ajax:
                    return jsonify({
                        "success": True,
                        "record_id": record_id,
                        "redirect_url": redirect_url,
                        "remaining_count": total_after_delete
                    })

                return redirect(redirect_url)

            if action == "import_json":
                json_text = (request.form.get("json_text") or "").strip()
                if not json_text:
                    raise ValueError("Vui lòng nhập JSON.")

                import_json_rows(dataset, fields, json_text)
                db.session.commit()
                return redirect(build_dataset_records_url(item=item, dataset=dataset, page=1, page_size=page_size))

            if action == "import_excel":
                excel_file = request.files.get("excel_file")
                if not excel_file or excel_file.filename == "":
                    raise ValueError("Vui lòng chọn file Excel.")

                import_excel_rows(dataset, fields, excel_file.read())
                db.session.commit()
                return redirect(build_dataset_records_url(item=item, dataset=dataset, page=1, page_size=page_size))

            raise ValueError("Thao tác không hợp lệ.")

        except Exception as e:
            db.session.rollback()
            if is_ajax:
                return jsonify({
                    "success": False,
                    "error": str(e)
                }), 400
            error = str(e)

    total_records = DanhMucRecord.query.filter_by(dataset_id=dataset.id).count()
    records_query = build_records_query(dataset.id, filter_conditions)
    filtered_total = records_query.count()
    total_pages = max(1, (filtered_total + page_size - 1) // page_size)

    if page > total_pages:
        page = total_pages
    if page < 1:
        page = 1

    records = (
        records_query
        .order_by(DanhMucRecord.id.asc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    record_rows = build_record_rows(records, fields)

    sync_unit = None
    try:
        sync_unit = resolve_sync_unit(item=item, dataset=dataset)
    except Exception:
        sync_unit = None

    scope = (item.scope or "COMMON").upper()
    if scope == "UNIT":
        sync_post_url = url_for("category_bp.sync_category_dataset", category_id=item.id, dataset_id=dataset.id)
        sync_login_url = url_for("category_bp.sync_category_dataset_login", category_id=item.id, dataset_id=dataset.id)
    else:
        sync_post_url = url_for("category_bp.sync_category_common", category_id=item.id)
        sync_login_url = url_for("category_bp.sync_category_common_login", category_id=item.id)

    return render_template(
        "category_records.html",
        item=item,
        dataset=dataset,
        fields=fields,
        records=records,
        record_rows=record_rows,
        error=error,
        sync_unit=sync_unit,
        sync_post_url=sync_post_url,
        sync_login_url=sync_login_url,
        page=page,
        page_size=page_size,
        total_pages=total_pages,
        total_records=total_records,
        filtered_total=filtered_total,
        page_start_index=((page - 1) * page_size) + 1 if filtered_total else 0,
        filter_conditions=filter_conditions
    )


def normalize_page_size(page_size):
    allowed_sizes = [10, 50, 200, 1000]
    if page_size not in allowed_sizes:
        return 50
    return page_size


def parse_filter_conditions(args):
    field_ids = args.getlist("filter_field_id")
    values = args.getlist("filter_value")

    conditions = []
    for field_id, value in zip(field_ids, values):
        field_id = to_int_or_none(field_id)
        value = (value or "").strip()
        if not field_id or not value:
            continue
        conditions.append({
            "field_id": field_id,
            "value": value
        })

    return conditions


def build_records_query(dataset_id, filter_conditions=None):
    query = DanhMucRecord.query.filter_by(dataset_id=dataset_id)
    filter_conditions = filter_conditions or []

    for condition in filter_conditions:
        field_id = condition.get("field_id")
        value = (condition.get("value") or "").strip()
        if not field_id or not value:
            continue

        matching_record_ids = (
            db.session.query(DanhMucRecordValue.record_id)
            .join(DanhMucRecord, DanhMucRecord.id == DanhMucRecordValue.record_id)
            .filter(DanhMucRecord.dataset_id == dataset_id)
            .filter(DanhMucRecordValue.field_id == field_id)
            .filter(func.coalesce(DanhMucRecordValue.value, "").ilike(f"%{value}%"))
        )
        query = query.filter(DanhMucRecord.id.in_(matching_record_ids))

    return query


def build_dataset_records_url(item, dataset, page=1, page_size=50, filter_conditions=None):
    scope = (item.scope or "COMMON").upper()
    if scope == "UNIT":
        base_url = url_for("category_bp.manage_category_dataset_records", category_id=item.id, dataset_id=dataset.id)
    else:
        base_url = url_for("category_bp.manage_category_records", category_id=item.id)

    params = [("page", page), ("page_size", normalize_page_size(page_size))]
    for condition in (filter_conditions or []):
        field_id = condition.get("field_id")
        value = (condition.get("value") or "").strip()
        if field_id and value:
            params.append(("filter_field_id", field_id))
            params.append(("filter_value", value))

    return base_url + "?" + urlencode(params, doseq=True)


def get_or_create_common_dataset(category):
    dataset = (
        DanhMucDataset.query
        .filter(DanhMucDataset.danh_muc_id == category.id)
        .filter(DanhMucDataset.don_vi_id.is_(None))
        .first()
    )
    if dataset:
        return dataset

    dataset = DanhMucDataset(
        danh_muc_id=category.id,
        don_vi_id=None,
        ten_bo_du_lieu="Bộ dữ liệu chung"
    )
    db.session.add(dataset)
    db.session.flush()
    return dataset


def import_json_rows(dataset, fields, json_text):
    data = json.loads(json_text)

    if isinstance(data, dict):
        if "rows" in data and isinstance(data["rows"], list):
            rows = data["rows"]
        else:
            rows = [data]
    elif isinstance(data, list):
        rows = data
    else:
        raise ValueError("JSON phải là object hoặc list object.")

    field_map = build_field_map(fields)

    for row in rows:
        if not isinstance(row, dict):
            continue

        record = DanhMucRecord(dataset_id=dataset.id)
        db.session.add(record)
        db.session.flush()

        for field in fields:
            raw_value = find_value_by_field(row, field, field_map)
            db.session.add(DanhMucRecordValue(
                record_id=record.id,
                field_id=field.id,
                value=normalize_cell_value(raw_value)
            ))


def import_excel_rows(dataset, fields, file_bytes, auto_create_fields=True):
    workbook = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("File Excel không có dữ liệu.")

    header_row_index = None
    stt_col_index = None

    for row_index, row in enumerate(rows):
        for col_index, cell_value in enumerate(row):
            cell_text = str(cell_value).strip().lower() if cell_value is not None else ""
            if cell_text == "stt":
                header_row_index = row_index
                stt_col_index = col_index
                break
        if header_row_index is not None:
            break

    if header_row_index is None or stt_col_index is None:
        raise ValueError("Không tìm thấy ô 'STT' để xác định bảng dữ liệu trong file Excel.")

    header_row = rows[header_row_index]
    headers = []
    for cell_value in header_row[stt_col_index:]:
        headers.append(str(cell_value).strip() if cell_value is not None else "")

    if not any(headers):
        raise ValueError("Không đọc được dòng tiêu đề của bảng Excel.")

    field_map = build_field_map(fields)

    if auto_create_fields:
        existing_keys = set(field_map.keys())

        for header in headers:
            header_text = (header or "").strip()
            header_key = header_text.lower()

            if not header_text:
                continue

            if header_key == "stt":
                continue

            if header_key not in existing_keys:
                new_field = DanhMucField(
                    danh_muc_id=dataset.danh_muc_id,
                    ma_truong=header_text
                )
                db.session.add(new_field)
                db.session.flush()

                fields.append(new_field)
                field_map[header_key] = new_field
                existing_keys.add(header_key)

    fields = (
        DanhMucField.query
        .filter_by(danh_muc_id=dataset.danh_muc_id)
        .order_by(DanhMucField.id.asc())
        .all()
    )
    field_map = build_field_map(fields)

    for data_row in rows[header_row_index + 1:]:
        row_dict = {}
        has_data = False

        for idx, header in enumerate(headers):
            if not header:
                continue

            source_col_index = stt_col_index + idx
            value = data_row[source_col_index] if source_col_index < len(data_row) else None

            if str(header).strip().lower() == "stt":
                continue

            if value is not None and str(value).strip() != "":
                has_data = True

            row_dict[header] = value

        if not has_data:
            continue

        record = DanhMucRecord(dataset_id=dataset.id)
        db.session.add(record)
        db.session.flush()

        for field in fields:
            raw_value = find_value_by_field(row_dict, field, field_map)
            db.session.add(DanhMucRecordValue(
                record_id=record.id,
                field_id=field.id,
                value=normalize_cell_value(raw_value)
            ))


def build_field_map(fields):
    result = {}
    for field in fields:
        result[(field.ma_truong or "").strip().lower()] = field
    return result


def find_value_by_field(row_dict, field, field_map):
    if not isinstance(row_dict, dict):
        return None

    ma_key = (field.ma_truong or "").strip()

    if ma_key in row_dict:
        return row_dict.get(ma_key)

    lowered = {}
    for k, v in row_dict.items():
        lowered[str(k).strip().lower()] = v

    if ma_key.lower() in lowered:
        return lowered.get(ma_key.lower())

    return None


def build_record_rows(records, fields):
    rows = []

    for record in records:
        value_map = {}
        for val in record.values:
            value_map[val.field_id] = val.value or ""

        row = {
            "record": record,
            "cells": []
        }

        for field in fields:
            row["cells"].append(value_map.get(field.id, ""))

        rows.append(row)

    return rows


def normalize_cell_value(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def to_int_or_none(value):
    if value is None:
        return None
    value = str(value).strip()
    return int(value) if value else None